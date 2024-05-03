# Job Application Automation
#
# @author Frankie Cook - github.com/frankiecook
# @version python3.12
# @since 0.0
# @date created 4/24/2024
#
# To-Do
# - undo temporary solution that doesn't certify ssl
# - add other websites (only works with LinkedIn)
# - create backup of save data
# - GUI addition
# - deeper website interaction (highlight selections, web browser plugin)
# - blacklist
# - safety fails for so many checks (missing sheet, missing data, url fail, missing excel file)
# - output successful save info (company, place, job, line)
# - limit retrys for url 
# - ampersand errors
#
# imports
from urllib import request
from datetime import date
import ssl
import openpyxl
import time
import os

###########
# VARIABLES
###########
file_name = 'ApplicationTracker.xlsx'
sheet = 'ATLA'
date_applied = date.today().strftime("%m/%d/%y")
search_location = 'Atlanta, GA'
search_terms = 'junior software developer'
save_file_name = "data.txt"
websites = ["LinkedIn","GlassDoor","Indeed","Monster","FlexJobs","ZipRecruiter","CareerBuilder","SimplyHired","Ladders","MetaCareers"]
website = ''

###########
# FUNCTIONS
###########
# check file path validity:
def doesFileExist(file_path):
    if not os.path.isfile(file_path):
        print("ERROR: ["+file_path+"] does not exist.")
        exit()

# checks if worksheet has data at given cell
# ws : openpyxl worksheet
# cell : cell location
def hasData(ws, cell):
    # check for data
    check_data = ws[cell].value

    if check_data != None:
        return True
    return False

# open and print request as html
#with request.urlopen(req) as f:
#    html = f.read().decode('utf-8')
def openHTML(req):
    try:
        f=request.urlopen(req)
        html_out = f.read().decode('utf-8')
        return html_out
    except:
        print('URL Failed. Waiting to Try Again.')
        time.sleep(1)   # seconds
        return openHTML(req)

# prompt for user input
def promptInput():
    usr_input = str(input("Enter URL: "))

    # special keywords
    if usr_input == "STOP" or usr_input == "stop" or usr_input=="S" or usr_input=="s":
        print("Goodbye Cowboy.")
        exit()
    
    # return
    return usr_input

# make request for the given url
def requestURL(url):
    # header tricks server
    try:
        req_out = request.Request(url, headers={'User-Agent': 'Mozilla'})
        return req_out
    except:
        print("Bad URL. Try Again.")
        return False

# load save data from txt file
def loadData(file_path):
    # check file path validity
    doesFileExist(file_path)

    # open file in read format
    file = open(file_path,'r')
    data = file.read()
    file.close()

    # split data into list
    data = data.replace('\n',',')
    data_list = data.split(",")

    # find starting row
    index = data_list.index(sheet) + 1
    row = int(data_list[index])
    return row

# create backup of file
def createBackup(file_path):
    backup_folder = "backups"
    backup_count = 1
    isExist = os.path.exists(backup_folder)

    # check if folder already exists
    if not isExist:
        os.mkdir(backup_folder)
        print("Backup Folder location created: "+backup_folder)

    # build path to back folder directory
    backup_path = str(backup_folder)+'/'+str(file_path)+'-bu'
    print(backup_path)

    # open workbook
    wb=openpyxl.load_workbook(file_name)
    # save file
    wb.save(backup_path)
    wb.close()
    print("Backup Created.")

# the removal of all occurrences of 
# a given item using filter() and __ne__ 
def remove_items(test_list, item): 
  
    # using filter() + __ne__ to perform the task 
    res = list(filter((item).__ne__, test_list)) 
    return res 

################
# READ SAVE FILE
################
# load data from save file
# for now, only row is kept track of
cur_row = loadData(save_file_name)

##################
# SCRAPE GIVEN URL
##################
# temporary solution towards failed certs
ssl._create_default_https_context = ssl._create_unverified_context

# check if spreadsheet file exists, create backup
doesFileExist(file_name)
createBackup(file_name)

######
# MAIN
######
def main():
    # variables
    global cur_row
    global company 
    global position
    global sheet

    # propmt use for URL
    # request abstraction of URL
    # open html of request
    url = promptInput()
    req = requestURL(url)
    # if request fails, then ask for new URL
    while not req:
        url = promptInput()
        req = requestURL(url)
    html = openHTML(req)

    # check url for website
    for site in websites:
        site_nocase = site.lower()
        inUrl = url.find(site_nocase)

        if inUrl != -1:
            website = site 
            print("Website is "+website)
            break
        
    # PARSE HTML
    ############
    # search html for values (company, position, actual location, search terms, link)
    offset = len('<title>')
    index_start = html.find('<title>') + offset
    index_end = html.find('</title>')

    # example of scrap
    # <company> hiring <position> in <actual location> | LinkedIn
    scrap = html[index_start:index_end]

    # company
    ci_start = 0
    ci_end = scrap.find(' hiring ')
    company = scrap[ci_start:ci_end]
    # position
    offset_pis = len(' hiring ')
    pi_start = scrap.find(' hiring ') + offset_pis
    pi_end = scrap.find(' in ')
    position = scrap[pi_start:pi_end]
    # actual location
    offset_alis = len(' in ')
    ali_start = scrap.find(' in') + offset_alis
    ali_end = scrap.find(' | ')
    actual_location = scrap[ali_start:ali_end]

    # WRITE RESULT TO SPREADSHEET
    #############################
    # variables
    row = str(cur_row)

    # open workbook and sheet
    # data_only : determines either the formula or last value stored
    wb=openpyxl.load_workbook(file_name, data_only=True)
    ws= wb[sheet]
    
    # spreadsheet cell : data
    application = {
        'A'+row:str(company),
        'B'+row:str(position),
        'C'+row:str(search_location),
        'D'+row:str(actual_location),
        'E'+row:str(date_applied),
        #'F'+row:str(),
        'G'+row:str(search_terms),
        'H'+row:str(url),
        'I'+row:str(website)
    }

    # check if application already exists in current sheet
    col_letters = ['A','B','D'] # F,H,I,E,C,G excluded
    matches = []
    
    # cycle through select columns
    for col_letter in col_letters:
        # find the current column
        column = ws[col_letter]

        # each column checks against different value
        check_application = application[str(col_letter+row)]

        # only save row value when match is found
        for col_cell in column:
            if col_cell.value == check_application:
                matches.append(col_cell.row)

    # check occurance of row values
    for match in matches:
        occurance = matches.count(match)
        if occurance == len(col_letters):
            matches = remove_items(matches,match)
            print("WARNING! Match Found at Row "+str(match))

            wb.close()
            exit()

    # check if row has any data for each cell
    for cell in application:
        # if row contains data, then exit program
        if hasData(ws, cell):
            print("ERROR: "+cell+" has data. Check saved row in data.txt")

            wb.close()
            exit()

    # save application to row
    for cell in application:
        ws[cell]=application[cell]

    # save file
    wb.save(file_name)
    wb.close()
    # output successful save
    print("\n~~APPLICATION SAVED~~\nsheet: {}\nrow: {}\ncompany: {}\nposition: {}\n".format(sheet,cur_row,company,position))

    # save current row
    cur_row+=1
    file = open("data.txt",'r')
    data_lines = file.readlines()
    file.close()

    for i in range(len(data_lines)):
        line = data_lines[i]

        # -1 if not found
        if line.find(sheet) != -1:
            # modify line, store current row
            new_line = sheet+","+str(cur_row)+"\n"
            data_lines[i] = new_line
    
    with open("data.txt",'w') as file:
        file.writelines(data_lines)
        file.close()

    main()

main()